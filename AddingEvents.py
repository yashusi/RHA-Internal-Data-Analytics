import tkinter as tk
from openpyxl import load_workbook
from tkinter import *
from tkcalendar import Calendar

wb = load_workbook("RHAexcelTest.xlsx")

#DICTIONARIES USED TO LOOK UP POINT VALUES
POINT_VALUES = {
"Early Proposal":10,
"Passive Program":30,
"Social":25,
"Educational":50,
"Social Justice":75,
"Traditional":100,

"Faculty/Pro Involvement":25,
"Community Collab 2":30,
"Community Collab 3":50,
"Community Collab 4":75,
"Community Collab 5":100,
"RA Collab":15,
"Outside Org":25,

"BYO":10,
"DIY Craft":10,
"Appreciation Event":40,

"Social Media Post":10,
"Email Listserv":5,
"Poster/Flyer":10,
"Paint Cube":50,
"HeelLife Post":15,

"Guest at BOG":15,
"OTM Submission":15,
"Community Wins OTM":100,
"Program OTM":50,
}

POPULATIONS = {
'Carmichael':400,
'Cobb':400,
'Conner':400,
'Craige':640,
'Eringhaus':640,
'Granville':500,
'Hinton James':950,
'Kenan':450,
'Quads':500,
'Mannings':550,
'Morrison':800,
'Rams':800,
}

#helper methods
def category_exists(category: str) -> bool:
    for row in POINT_VALUES:
        if (str(category) == str(row[0])):
            return True
    return False

def point_value_lookup(category:str) -> int:
    return POINT_VALUES[category]

def points_calculator(categories: list) -> int:
    total: int = 0
    for item in categories:
        total += int(point_value_lookup(item))
    return total

def attendance_to_points(community: str, att: int) -> int:
    if(att == 0 or att == -1):
        return 0
    percent: float = 0.0
    percent = (float)(att)/(POPULATIONS[community])
    
    if(percent>.57):
        return 75
    if(percent>.42):
        return 50
    if(percent>.33):
        return 35
    if(percent>.12):
        return 20
    print("10 points were added from attandence")
    return 10

def list_to_string(categories_list: list[str]) -> str:
    categories_string = ""
    for item in categories_list:
        categories_string += item
        categories_string += ", "
    return categories_string

def adding_proposal_to_excel():
    #important fields
    community: str
    event_name: str
    date: str
    attendance = 0
    point_total: int
    list_of_categories = []

    #defining tkinter window
    window = tk.Tk()
    window.title("Fill out the following fields")
    window.geometry("800x500")

    #frames to organize GUI horizantally as opposed to vertically (default)
    date_name_frame = LabelFrame(window, text="Select Event Name and Date", padx = 5, pady = 5)
    date_name_frame.pack(side = LEFT)

    categories_frame = LabelFrame(window, text="Select Categories", padx = 5, pady = 5)
    categories_frame.pack(side = LEFT)

    community_frame = LabelFrame(window, text="Select Community", padx = 5, pady = 5)
    community_frame.pack(side = LEFT)

    #Event name textbox
    def printInput():
        inp = inputtxt.get(1.0, "end-1c")
    # TextBox Creation
    inputtxt = tk.Text(date_name_frame, height = 1, width = 20)
    inputtxt.pack()
    
    #calendar for event date
    cal = Calendar(date_name_frame, selectmode = 'day',
               year = 2023, month = 1,
               day = 1)
    cal.pack()
 
    #dropdown menu for chossing community
    community_options = [
        'Carmichael',
        'Cobb',
        'Conner',
        'Craige',
        'Eringhaus',
        'Granville',
        'Hinton James',
        'Kenan',
        'Quads',
        'Mannings',
        'Morrison',
        'Rams',
    ]
    clicked = StringVar()
    clicked.set(community_options[0])
    drop = OptionMenu(window, clicked, *community_options)
    drop.pack()


    #l = tk.Label(window, bg="white", width=20, text="empty")
    #l.pack(side = RIGHT)

    def print_selection():
        if(enter.get() == 1):
            #l.config(text='done')
            event_name = inputtxt.get(1.0, "end-1c")
            date = cal.get_date()
            community = clicked.get()
            point_total = points_calculator(list_of_categories) + (int)(attendance_to_points(community, attendance))
            sheet = wb[community]
            last_row = sheet.max_row
            while(sheet.cell(last_row, 1).value is None):
                last_row -= 1
            next_row = last_row+1
            data = [date, event_name, list_to_string(list_of_categories), attendance, point_total]
            for col, value in enumerate(data, start=1):
                sheet.cell(row=next_row, column=col).value = value
            #sheet.append([date, event_name, list_to_string(category_list), attendance, point_total])
            wb.save("RHAexcelTest.xlsx")
            window.destroy()
            
        else:
            if(early.get() == 1):
                #l.config(text="Early Proposal")
                list_of_categories.append("Early Proposal")
                early.set(0)
                C_early["bg"] = "green"
                
            if(social.get() == 1):
                #l.config(text="SOCIAL")
                list_of_categories.append("Social")
                social.set(0)
                C_social["bg"] = "green"

            if(edu.get() == 1):
                #l.config(text='educational')
                list_of_categories.append("Educational")
                edu.set(0)
                C_edu["bg"] = "green"

            if(socjus.get() == 1):
                #l.config(text='Social Justice')
                list_of_categories.append("Social Justice")
                socjus.set(0)
                C_socjus["bg"] = "green"

            if(trad.get() == 1):
                #l.config(text='Traditional')
                list_of_categories.append("Traditional")
                trad.set(0)
                C_trad["bg"] = "green"

            if(ra_collab.get() == 1):
                #l.config(text='RA Collab')
                list_of_categories.append("RA Collab")
                ra_collab.set(0)
                C_ra_collab["bg"] = "green"
            
            if(co_collab2.get() == 1):
                #l.config(text='Community Collab: 2')
                list_of_categories.append("Community Collab: 2")
                co_collab2.set(0)
                C_co_collab2["bg"] = "green"
            
            if(byo.get() == 1):
                #l.config(text='BYO')
                list_of_categories.append("BYO")
                byo.set(0)
                C_byo["bg"] = "green"

            if(diy.get() == 1):
                #l.config(text='DIY Craft')
                list_of_categories.append("DIY Craft")
                diy.set(0)
                C_diy["bg"] = "green"
            
            if(passive.get() == 1):
                #l.config(text='Passive Program')
                list_of_categories.append("Passive Program")
                passive.set(0)
                C_passive["bg"] = "green"

            if(smpost.get() == 1):
                #l.config(text='Social Media Post')
                list_of_categories.append("Social Media Post")
                smpost.set(0)
                C_smpost["bg"] = "green"

            if(flyer.get() == 1):
                #l.config(text='Poster/Flyer')
                list_of_categories.append("Poster/Flyer")
                flyer.set(0)
                C_flyer["bg"] = "green"

            if(heel_life.get() == 1):
                #l.config(text='Heel Life Post')
                list_of_categories.append("HeelLife Post")
                heel_life.set(0)
                C_heel_life["bg"] = "green"

            if(email.get() == 1):
                #l.config(text='Email Listserv')
                list_of_categories.append("Email Listserv")
                email.set(0)
                C_email["bg"] = "green"
            
            if(bog_guest.get() == 1):
                #l.config(text='Guest at BOG')
                list_of_categories.append("Guest at BOG")
                bog_guest.set(0)
                C_bog_guest["bg"] = "green"
        

    enter = tk.IntVar()
    early = tk.IntVar()

    social = tk.IntVar()
    edu = tk.IntVar()
    socjus = tk.IntVar()
    trad = tk.IntVar()

    ra_collab = tk.IntVar()
    co_collab2 = tk.IntVar()

    byo = tk.IntVar()
    diy = tk.IntVar()

    passive = tk.IntVar()

    smpost = tk.IntVar()
    flyer = tk.IntVar()
    heel_life = tk.IntVar()
    email = tk.IntVar()

    bog_guest = tk.IntVar()




    #Common Categories
    C_enter = tk.Checkbutton(window, text='Enter',variable=enter, onvalue=1, offvalue=0, command=print_selection)
    C_enter.pack()

    C_early = tk.Checkbutton(window, text='Early Proposal',variable=early, onvalue=1, offvalue=0, command=print_selection)
    C_early.pack()
    #event types
    C_social = tk.Checkbutton(window, text='Social',variable=social, onvalue=1, offvalue=0, command=print_selection)
    C_social.pack()

    C_edu = tk.Checkbutton(window, text='Educational',variable=edu, onvalue=1, offvalue=0, command=print_selection)
    C_edu.pack()

    C_socjus = tk.Checkbutton(window, text='Social Justice',variable=socjus, onvalue=1, offvalue=0, command=print_selection)
    C_socjus.pack()

    C_trad = tk.Checkbutton(window, text='Traditional',variable=trad, onvalue=1, offvalue=0, command=print_selection)
    C_trad.pack()

    #collabs
    C_ra_collab = tk.Checkbutton(window, text='RA Collab',variable=ra_collab, onvalue=1, offvalue=0, command=print_selection)
    C_ra_collab.pack()

    C_co_collab2 = tk.Checkbutton(window, text='Commnunity Collab: 2',variable=co_collab2, onvalue=1, offvalue=0, command=print_selection)
    C_co_collab2.pack()

    #event attributes
    C_byo = tk.Checkbutton(window, text='BYO',variable=byo, onvalue=1, offvalue=0, command=print_selection)
    C_byo.pack()

    C_diy = tk.Checkbutton(window, text='DIY Craft',variable=diy, onvalue=1, offvalue=0, command=print_selection)
    C_diy.pack()


    C_passive = tk.Checkbutton(window, text='Passive Program',variable=passive, onvalue=1, offvalue=0, command=print_selection)
    C_passive.pack()
    #marketting
    C_smpost = tk.Checkbutton(window, text='Social Media Post',variable=smpost, onvalue=1, offvalue=0, command=print_selection)
    C_smpost.pack()

    C_flyer = tk.Checkbutton(window, text='Poster/Flyer',variable=flyer, onvalue=1, offvalue=0, command=print_selection)
    C_flyer.pack()

    C_heel_life = tk.Checkbutton(window, text='HeelLife Post',variable=heel_life, onvalue=1, offvalue=0, command=print_selection)
    C_heel_life.pack()

    C_email = tk.Checkbutton(window, text='Email Listserv',variable=email, onvalue=1, offvalue=0, command=print_selection)
    C_email.pack()

    C_bog_guest = tk.Checkbutton(window, text='Guest at BOG',variable=bog_guest, onvalue=1, offvalue=0, command=print_selection)
    C_bog_guest.pack()
    
    window.mainloop()
adding_proposal_to_excel()